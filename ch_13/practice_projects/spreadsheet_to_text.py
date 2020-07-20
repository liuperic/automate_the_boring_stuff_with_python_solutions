#!/usr/bin/env python3

"""Copies contents of spreadsheet columns into text files. Takes command line arguments
    files MUST be in current working directory.
    Usage: python spreadsheet_to_text.py <Excel file>"""

import openpyxl, sys
from openpyxl.styles import Font

if len(sys.argv) != 2:
    print('Usage error: python text_files.py <Excel file>\n' + 'Please enter the name of the excel file you would like to read fron.')

wb_name = sys.argv[1]

wb = openpyxl.load_workbook(wb_name)
sheet = wb.active

file_num = 1

for column_num in range(1, sheet.max_column + 1):
    # Name of text file
    txt_file =  'text' + str(file_num) + '.txt'
    with open(txt_file, 'w') as file:
        for row_num in range(1, sheet.max_row + 1):
            # Read each row from excel if cell not empty
            if sheet.cell(row=row_num, column=column_num).value != None:
                # Strip out any new lines at end of each row content
                line = str(sheet.cell(row=row_num, column=column_num).value).rstrip()
                file.write(line + '\n')
    # New file per new column
    file_num += 1

# Each column in excel will be saved into 'text1.txt' where the number iterates for each new file.
print(f'Contents of each column in {wb_name} has been saved into text files in current directory.')
       
