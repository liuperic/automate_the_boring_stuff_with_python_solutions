#!/usr/bin/env python3
# multiplication_table.py - Creates a multiplication table on excel
# NxN multiplication table based on input from command line

import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

n = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
font_bold_obj = Font(bold=True)

# Create column for multiplication table
for count, row_num in enumerate(range(2, n + 2), 1):    # Start 2nd row and continue for n + 1 elements
    sheet['A' + str(row_num)].font = font_bold_obj
    sheet['A' + str(row_num)] = count
wb.save('multiplication_table.xlsx')

# Create row for multiplication table
for count, column_num in enumerate(range(2, n + 2), 1):    # Start 2nd row and continue for n + 1 elements
    sheet.cell(row=1, column=column_num).font = font_bold_obj
    sheet.cell(row=1, column=column_num).value = count

# Iterate through multiplication table and fill out
for row_num in range(2, n + 2):
    for column_num in range(2, n + 2):
        column_table = get_column_letter(column_num)
        # Create excel formula to get product of multiplication table for each cell
        cell_content = '=' + column_table + '1*A' + str(row_num)
        sheet.cell(row=row_num, column=column_num).value = cell_content
wb.save('multiplication_table.xlsx')
