#!/usr/bin/env python3

import openpyxl, logging
from openpyxl.utils import get_column_letter

def transpose(file_name, output_file):
    """transpose inverts row and columns of cells in spreadsheet

    Args:
        file_name (string): name of workbook to transpose data
        output_file (string): name of workbook to save transposed data in
    Summary:
        Will iterate through all sheets in file_name and tranpose data into
        corresponding sheets in output_file
    """
    # Currently disabled logging messages
    logging.disable(logging.CRITICAL)

    logging.debug(f'Start of tranpose({file_name, output_file})')
    wb = openpyxl.load_workbook(file_name)
    # Open new workbook
    new_wb = openpyxl.Workbook()
    # Save default title of active sheet and delete later
    act_sheet = new_wb.active.title

    # Loop through sheets
    for sheet in wb.sheetnames:
        # Create a new sheet for each sheet
        new_wb.create_sheet(title=sheet)
        current_sheet = new_wb[sheet]
        logging.info('Current sheet is ' + sheet)
        for row_num in range(1, wb[sheet].max_row + 1):
            for column_num in range(1, wb[sheet].max_column + 1):
                cell_content = wb[sheet].cell(row=row_num, column=column_num).value
                logging.debug(f'Cell row: {row_num}, column: {get_column_letter(column_num)}; content contains: {cell_content}')
                current_sheet.cell(row=column_num, column=row_num).value = cell_content
                logging.debug(f'Copying to contents to new sheet in Cell row: {column_num}, column: {get_column_letter(row_num)}; value of cell: {current_sheet.cell(row=column_num, column=row_num).value}')
    # Delete default active sheet
    del new_wb[act_sheet]
    logging.info(f'Sheet names: {new_wb.sheetnames}')
    # Save file
    new_wb.save(output_file)


if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
    transpose('example.xlsx', 'example2.xlsx')
    logging.shutdown()