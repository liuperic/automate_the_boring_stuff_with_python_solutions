#!/usr/bin/env python3

"""Copies contents of 1 or more text files into a spreadsheet. Takes command line arguments
    files MUST be in current working directory.
    Usage: python text_files.py <Excel file> <text file #1> <text file #2> ..."""

import openpyxl, sys
from openpyxl.styles import Font

if len(sys.argv) < 3:
    print('Usage error: python text_files.py <Excel file> <textfile>\n' + 'Please enter the name of the excel file you would like to save the text files in and enter at least one text file')

wb_name = sys.argv[1]
files = sys.argv[2:]

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'txt'

column_num = 1

for txt_file in files:
    # Read each text file line by line
    lines = open(txt_file).readlines()
    
    bold = Font(bold=True)
    sheet.cell(row=1, column=column_num).value = txt_file
    sheet.cell(row=1, column=column_num).font = bold
    row_num = 2
    # Save each line in a cell
    for line in lines:
        sheet.cell(row=row_num, column=column_num).value = line
        row_num += 1
    column_num += 1

wb.save(wb_name)

print(f'files: {", ".join(files)} have been saved into {wb_name}')
       
